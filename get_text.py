# Most of code I got it from http://www.blog.pythonlibrary.org/2018/05/03/exporting-data-from-pdfs-with-python/
# and adapted it for my needs. Credits entriely go to Mike.

import io
import openpyxl

from pdfminer.converter import TextConverter
from pdfminer.pdfinterp import PDFPageInterpreter
from pdfminer.pdfinterp import PDFResourceManager
from pdfminer.pdfpage import PDFPage

file = 'data.xlsx'
wb = openpyxl.load_workbook(file)
sheet = wb['Sheet1']


def extract_text_from_pdf(pdf_path):
    resource_manager = PDFResourceManager()
    fake_file_handle = io.StringIO()
    converter = TextConverter(resource_manager, fake_file_handle)
    page_interpreter = PDFPageInterpreter(resource_manager, converter)

    with open(pdf_path, 'rb') as fh:
        for page in PDFPage.get_pages(fh, caching=True, check_extractable=True):
            page_interpreter.process_page(page)

        text = fake_file_handle.getvalue()

    converter.close()
    fake_file_handle.close()
    if text:
        return text


if __name__ == '__main__':
    text = extract_text_from_pdf('directory.pdf').split("Empresa Española:")

    sheet.cell(row=1, column=1).value = 'Empresa'
    sheet.cell(row=1, column=2).value = 'Dirección'
    sheet.cell(row=1, column=3).value = 'C.P. / Ciudad'
    sheet.cell(row=1, column=4).value = 'Estado / Provincia'
    sheet.cell(row=1, column=5).value = 'Zona'
    sheet.cell(row=1, column=6).value = 'Sectores sede local'
    sheet.cell(row=1, column=7).value = 'Telefonos'
    sheet.cell(row=1, column=8).value = 'Fax'
    sheet.cell(row=1, column=9).value = 'Email'
    sheet.cell(row=1, column=10).value = 'Web'
    # for i in range(0, 4):
    #     print("\n", i, "\n", text[i])

    for i in range(0, 615):
        try:
            name = text[i + 1].split('Dirección')[0]
            sheet.cell(row=i + 2, column=1).value = str(name)

            direccion = text[i].split('Dirección')[1].split('C.P. / Ciudad')[0]
            sheet.cell(row=i + 2, column=2).value = str(direccion)

            ciudad = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[0]
            sheet.cell(row=i + 2, column=3).value = str(ciudad)

            estado = \
                text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                    "Zona")[0]
            sheet.cell(row=i + 2, column=4).value = str(estado)

            zona = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                "Zona")[
                1].split(
                'Sectores sede local')[0]
            sheet.cell(row=i + 2, column=5).value = str(zona)

            sectores = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                "Zona")[
                1].split(
                'Sectores sede local')[1].split("Teléfono")[0]
            sheet.cell(row=i + 2, column=6).value = str(sectores)

            telefono = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                "Zona")[1].split('Sectores sede local')[1].split("Teléfono")[1].split("Fax")[0]
            sheet.cell(row=i + 2, column=7).value = str(telefono)

            fax = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                "Zona")[1].split('Sectores sede local')[1].split("Teléfono")[1].split("Fax")[1].split("Email")[0]
            sheet.cell(row=i + 2, column=8).value = str(fax)

            email = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                "Zona")[1].split(
                'Sectores sede local')[1].split("Teléfono")[1].split("Fax")[1].split("Email")[1].split("Web")[0]
            sheet.cell(row=i + 2, column=9).value = str(email)

            web = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                "Zona")[1].split(
                'Sectores sede local')[1].split("Teléfono")[1].split("Fax")[1].split("Email")[1].split("Web")[
                1].split(" ")[0]
            sheet.cell(row=i + 2, column=10).value = str(web)

            wb.save(file)
        except:
            try:
                name = text[i + 1].split('SA')[0]
                name = name + 'SA'
                sheet.cell(row=i + 2, column=1).value = str(name)

                direccion = text[i].split('Dirección')[1].split('C.P. / Ciudad')[0]
                sheet.cell(row=i + 2, column=2).value = str(direccion)

                ciudad = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[0]
                sheet.cell(row=i + 2, column=3).value = str(ciudad)

                estado = \
                    text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                        "Zona")[0]
                sheet.cell(row=i + 2, column=4).value = str(estado)

                zona = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                    "Zona")[
                    1].split(
                    'Sectores sede local')[0]
                sheet.cell(row=i + 2, column=5).value = str(zona)

                sectores = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                    "Zona")[
                    1].split(
                    'Sectores sede local')[1].split("Teléfono")[0]
                sheet.cell(row=i + 2, column=6).value = str(sectores)

                telefono = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                    "Zona")[1].split('Sectores sede local')[1].split("Teléfono")[1].split("Fax")[0]
                sheet.cell(row=i + 2, column=7).value = str(telefono)

                fax = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                    "Zona")[1].split('Sectores sede local')[1].split("Teléfono")[1].split("Fax")[1].split("Email")[0]
                sheet.cell(row=i + 2, column=8).value = str(fax)

                email = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                    "Zona")[1].split(
                    'Sectores sede local')[1].split("Teléfono")[1].split("Fax")[1].split("Email")[1].split("Web")[0]
                sheet.cell(row=i + 2, column=9).value = str(email)

                web = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                    "Zona")[1].split(
                    'Sectores sede local')[1].split("Teléfono")[1].split("Fax")[1].split("Email")[1].split("Web")[
                    1].split(" ")[0]
                sheet.cell(row=i + 2, column=10).value = str(web)

                wb.save(file)

            except:
                try:
                    name = text[i + 1].split('Página')[0]
                    sheet.cell(row=i + 2, column=1).value = str(name)

                    direccion = text[i].split('Dirección')[1].split('C.P. / Ciudad')[0]
                    sheet.cell(row=i + 2, column=2).value = str(direccion)

                    ciudad = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[0]
                    sheet.cell(row=i + 2, column=3).value = str(ciudad)

                    estado = \
                        text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                            "Zona")[0]
                    sheet.cell(row=i + 2, column=4).value = str(estado)

                    zona = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                        "Zona")[
                        1].split(
                        'Sectores sede local')[0]
                    sheet.cell(row=i + 2, column=5).value = str(zona)

                    sectores = \
                        text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                            "Zona")[
                            1].split(
                            'Sectores sede local')[1].split("Teléfono")[0]
                    sheet.cell(row=i + 2, column=6).value = str(sectores)

                    telefono = \
                        text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                            "Zona")[1].split('Sectores sede local')[1].split("Teléfono")[1].split("Fax")[0]
                    sheet.cell(row=i + 2, column=7).value = str(telefono)

                    fax = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                        "Zona")[1].split('Sectores sede local')[1].split("Teléfono")[1].split("Fax")[1].split("Email")[
                        0]
                    sheet.cell(row=i + 2, column=8).value = str(fax)

                    email = \
                        text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                            "Zona")[1].split(
                            'Sectores sede local')[1].split("Teléfono")[1].split("Fax")[1].split("Email")[1].split(
                            "Web")[0]
                    sheet.cell(row=i + 2, column=9).value = str(email)

                    web = text[i].split('Dirección')[1].split('C.P. / Ciudad')[1].split("Estado / Provincia")[1].split(
                        "Zona")[1].split(
                        'Sectores sede local')[1].split("Teléfono")[1].split("Fax")[1].split("Email")[1].split("Web")[
                        1].split(" ")[0]
                    sheet.cell(row=i + 2, column=10).value = str(web)

                except:
                    print("\nSome error in i = ", i)
                    print("i", text[i])
